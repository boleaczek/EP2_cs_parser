from enum import Enum
from openpyxl import load_workbook
import json
import sys

class GeneratedFromExcel:
    def read_from_excel(self, sheet):
        for value_name, layout_type in self.values_layout.items():
            value = layout_type.read(sheet)
            self.add_value(value_name, value)

class DataLayout:
    def read(self, sheet):
        return self.read_data(sheet)

class SingleCell(DataLayout):
    def __init__(self, cell_id):
        self.cell_id = cell_id
    
    def read_data(self, sheet):
        return sheet[self.cell_id].value

class RowRange(DataLayout):
    def __init__(self, col, start, stop):
        self.col = col
        self.start = start
        self.stop = stop
    
    def read_data(self, sheet):
        data = ""
        for row in range(self.start, self.stop + 1):
            data += sheet[self.col + str(row)].value + " "
        return data

class ColRange(DataLayout):
    def __init__(self, row, start, stop):
            self.row = row
            self.start = ord(start)
            self.stop = ord(stop)

    def read_data(self, sheet):
        for col in range(self.start, self.stop):
            print("freading from {}", chr(col) + str(self.row))
        return None

class DataChunk(GeneratedFromExcel):
    values_layout = {}
    values = {}
    def add_value(self, val_name, value):
        self.values[val_name] = value

    def __init__(self, schema):
        self.init_from_schema(schema)

    def init_from_schema(self, schema):
        for schema_entry, layout in schema.items():
            if isinstance(layout, dict):
                if "col" in layout:
                    self.values_layout[schema_entry] = RowRange(layout["col"], layout["row_start"], layout["row_stop"])
                if "row" in layout:
                    self.values_layout[schema_entry] = RowRange(layout["row"], layout["col_start"], layout["col_stop"])
            else:
                self.values_layout[schema_entry] = SingleCell(layout)

character_sheet = sys.argv[1]
output = sys.argv[2]
schema = "schema.json"
if len(sys.argv) == 4:
    schema = sys.argv[3]

wb = load_workbook(filename = 'test.xlsx')
sheets = wb.sheetnames
ws = wb[sheets[0]]

with open(schema) as s:
    data = json.load(s)
    pd = DataChunk(data)
    pd.read_from_excel(ws)
    
    with open(output, 'w') as o:
        json.dump(pd.values, o, indent = 4)
    